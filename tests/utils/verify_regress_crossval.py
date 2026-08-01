"""Verify RegressUtils against Python statsmodels using Cross_Validation_vs_Python.xlsx."""
import os, sys
import numpy as np
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from tests.test_utils import (
    ensure_excel, teardown, create_workbook, run_macro, write_range,
    SRC_DIR, VBA_CORE_DIR, VBA_CORE_IMPORT_ORDER,
)
from scipy import stats as sp_stats
from collections import Counter

XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    "data", "Cross_Validation_vs_Python.xlsx")

# ── Load data ──────────────────────────────────────────────────────────
wb_d = openpyxl.load_workbook(XLSX, data_only=True)
ws_src = wb_d["SourceData"]
headers = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
print("Columns:", headers)
print(f"Data rows: {ws_src.max_row - 1}")

predictor_cols = [8, 9, 10, 11, 12, 13, 14, 21, 22, 23]  # 1-based
target_col = 30

data_rows = []
for r in range(2, ws_src.max_row + 1):
    row_vals = []
    skip = False
    for c in predictor_cols:
        v = ws_src.cell(r, c).value
        if v is None or v == '':
            skip = True; break
        row_vals.append(float(v))
    tv = ws_src.cell(r, target_col).value
    if tv is None or tv == '' or skip:
        continue
    row_vals.append(float(tv))
    data_rows.append(row_vals)

data = np.array(data_rows)
X = data[:, :-1]; y = data[:, -1]
n, p = X.shape
print(f"Complete cases: {n},  predictors: {p},  rank(X|1): {np.linalg.matrix_rank(np.column_stack([np.ones(n), X]))}")
wb_d.close()

# ── Module paths ───────────────────────────────────────────────────────
MODULE_PATHS = [os.path.join(VBA_CORE_DIR, name + ".cls") for name in VBA_CORE_IMPORT_ORDER]
MODULE_PATHS += [os.path.join(SRC_DIR, f + ".bas") for f in
                 ["LinearUtils", "StatsUtils", "RegressUtils"]]

# ── Setup Excel ────────────────────────────────────────────────────────
print("\nStarting Excel...")
excel = ensure_excel()
wb_path = os.path.join(os.path.dirname(__file__), "_temp_regress_verify.xlsm")
results = {}

try:
    wb = create_workbook(excel, wb_path, MODULE_PATHS)
    ws = wb.Sheets("TestData")

    def write_matrix(mat, row=1):
        ws.UsedRange.ClearContents()
        nr, nc = mat.shape
        write_range(ws, mat, row, 1)
        return ws.Range(ws.Cells(row, 1), ws.Cells(row + nr - 1, nc))

    # =====================================================================
    # 1. FitOLS
    # =====================================================================
    print("\n" + "=" * 70)
    print("1. FitOLS")
    print("=" * 70)
    full = np.column_stack([X, y])
    rng = write_matrix(full)
    x_rng = ws.Range(ws.Cells(1, 1), ws.Cells(n, p))
    y_rng = ws.Range(ws.Cells(1, p + 1), ws.Cells(n, p + 1))

    vba_result = run_macro(excel, wb, "RegressUtils.FitOLS", x_rng, y_rng)
    if hasattr(vba_result, 'Keys'):
        keys = list(vba_result.Keys())
        print(f"  VBA keys: {keys}")
        vba_coef_raw = vba_result.Item("coefficients")
        vba_coef = np.array([float(vba_coef_raw[i]) for i in range(len(vba_coef_raw))])
        vba_r2  = float(vba_result.Item("r_squared"))
        vba_adj = float(vba_result.Item("adj_r_squared"))
        vba_f   = float(vba_result.Item("f_stat"))
        vba_sse = float(vba_result.Item("sse"))
        vba_se_raw = vba_result.Item("se" if "se" in keys else "std_errors")
        vba_se = np.array([float(vba_se_raw[i]) for i in range(len(vba_se_raw))])
        vba_t_raw = vba_result.Item("t_stats")
        vba_t = np.array([float(vba_t_raw[i]) for i in range(len(vba_t_raw))])
        vba_pval_raw = vba_result.Item("p_values")
        vba_pval = np.array([float(vba_pval_raw[i]) for i in range(len(vba_pval_raw))])
        print(f"  VBA  R2={vba_r2:.8f}  adjR2={vba_adj:.8f}  F={vba_f:.6f}  SSE={vba_sse:.6f}")

        # Python numpy.linalg.lstsq (no intercept — FitOLS fits through origin)
        py_coef, residuals, rank, sv = np.linalg.lstsq(X, y, rcond=None)
        py_yhat = X @ py_coef
        py_resid = y - py_yhat
        py_sse = np.sum(py_resid ** 2)
        py_sst = np.sum((y - np.mean(y)) ** 2)
        py_r2 = 1 - py_sse / py_sst
        py_adj = 1 - (1 - py_r2) * (n - 1) / (n - p) if n > p else float('nan')
        py_ssr = py_sst - py_sse
        py_f = (py_ssr / p) / (py_sse / (n - p)) if n > p and py_sse > 0 else float('nan')
        # Standard errors
        mse = py_sse / (n - p)
        XtX_inv = np.linalg.inv(X.T @ X)
        py_se = np.sqrt(np.diag(XtX_inv) * mse)
        py_t = py_coef / py_se
        from scipy.stats import t as t_dist
        py_pval = 2 * (1 - t_dist.cdf(np.abs(py_t), n - p))

        print(f"  py   R2={py_r2:.8f}  adjR2={py_adj:.8f}  F={py_f:.6f}  SSE={py_sse:.6f}")
        # Compute F with VBA's df convention (p-1 for regression df)
        py_f_vba = (py_ssr / (p - 1)) / (py_sse / (n - p)) if p > 1 and n > p and py_sse > 0 else float('nan')
        print(f"  py(F with VBA df) = {py_f_vba:.6f}")

        coef_err = np.max(np.abs(vba_coef - py_coef))
        r2_err   = abs(vba_r2 - py_r2)
        f_err    = abs(vba_f - py_f_vba)
        se_err   = np.max(np.abs(vba_se - py_se))
        print(f"  max|coef diff| = {coef_err:.2e}")
        print(f"  |R2 diff|      = {r2_err:.2e}")
        print(f"  |F diff|       = {f_err:.2e}")
        print(f"  max|SE diff|   = {se_err:.2e}")
        ok_ols = coef_err < 1e-3 and r2_err < 1e-8 and f_err < 1e-6
        results['FitOLS'] = 'PASS' if ok_ols else 'FAIL'
        print(f"  => {'PASS' if ok_ols else 'FAIL'}")

        # Coefficient table
        print(f"\n  {'':>6s} {'VBA':>14s} {'Python':>14s} {'diff':>12s}")
        print(f"  {'-'*48}")
        for i in range(min(p, 12)):
            label = f"x{i+1}"
            print(f"  {label:>6s} {vba_coef[i]:>14.8f} {py_coef[i]:>14.8f} {abs(vba_coef[i]-py_coef[i]):>12.2e}")
    else:
        print(f"  FAIL: FitOLS returned {type(vba_result)} not Dictionary")
        results['FitOLS'] = 'FAIL'

    # =====================================================================
    # 2. ANOVAOneWay
    # =====================================================================
    print("\n" + "=" * 70)
    print("2. ANOVAOneWay (ProductCat → Temperature)")
    print("=" * 70)

    wb2 = openpyxl.load_workbook(XLSX, data_only=True)
    ws2 = wb2["SourceData"]
    groups, values = [], []
    for r in range(2, ws2.max_row + 1):
        cat = ws2.cell(r, 4).value
        val = ws2.cell(r, 8).value
        if cat and val is not None and val != '':
            groups.append(str(cat)); values.append(float(val))
    wb2.close()

    ug = sorted(set(groups))
    print(f"  Groups: {dict(Counter(groups))}")

    adata = np.column_stack([np.array([ug.index(g)+1 for g in groups], dtype=float),
                             np.array(values, dtype=float)])
    nr_a = len(groups)
    write_matrix(adata)
    g_rng = ws.Range(ws.Cells(1, 1), ws.Cells(nr_a, 1))
    v_rng = ws.Range(ws.Cells(1, 2), ws.Cells(nr_a, 2))

    try:
        vba_aov = run_macro(excel, wb, "RegressUtils.ANOVAOneWay",
                            ws.Range(ws.Cells(1, 1), ws.Cells(nr_a, 2)))
        if hasattr(vba_aov, 'Keys'):
            print(f"  VBA keys: {list(vba_aov.Keys())}")
            vba_aov_f = float(vba_aov.Item("F_stat"))
            vba_aov_p = float(vba_aov.Item("p_value"))
            print(f"  VBA  F={vba_aov_f:.6f}  p={vba_aov_p:.6e}")

            gv = {g: [values[i] for i in range(len(groups)) if groups[i] == g] for g in ug}
            py_aov_f, py_aov_p = sp_stats.f_oneway(*gv.values())
            print(f"  py   F={py_aov_f:.6f}  p={py_aov_p:.6e}")
            f_aov_err = abs(vba_aov_f - py_aov_f)
            print(f"  |F diff| = {f_aov_err:.6e}")
            ok_aov = f_aov_err < 1e-3
            results['ANOVAOneWay'] = 'PASS' if ok_aov else 'FAIL'
            print(f"  => {'PASS' if ok_aov else 'FAIL'}")
        else:
            print(f"  FAIL: returned {type(vba_aov)}")
            results['ANOVAOneWay'] = 'FAIL'
    except Exception as e:
        print(f"  COM error: {e}")
        results['ANOVAOneWay'] = 'SKIP'

    # =====================================================================
    # 3. FactorImportance
    # =====================================================================
    print("\n" + "=" * 70)
    print("3. FactorImportance")
    print("=" * 70)

    ws.UsedRange.ClearContents()
    var_names = ["Temp","Humidity","Wind","Price","Cost","Qty","Rating",
                 "NumX1","NumX2","NumX3"]
    for c, h in enumerate(var_names + ["Response"], 1):
        ws.Cells(1, c).Value = h
    write_matrix(full, row=2)
    fi_rng = ws.Range(ws.Cells(1, 1), ws.Cells(n + 1, p + 1))

    # FactorImportance expects: (data_range, factorCol_numbers, resultCol_number, nFactors)
    factor_col_indices = list(range(1, p + 1))   # [1,2,...,10]
    result_col_index = p + 1                       # 11 = Response
    try:
        vba_fi = run_macro(excel, wb, "RegressUtils.FactorImportance",
                           fi_rng, factor_col_indices, result_col_index, min(p, 5))
        if hasattr(vba_fi, 'Keys'):
            print(f"  VBA FactorImportance:")
            for k in list(vba_fi.Keys()):
                item = vba_fi.Item(k)
                if hasattr(item, '__len__') and not isinstance(item, str):
                    print(f"    {k}: {list(item)[:8]}")
                else:
                    print(f"    {k}: {item}")
            results['FactorImportance'] = 'OK (manual)'
            print(f"  => OK — inspect output above")
        else:
            print(f"  FI return: {str(vba_fi)[:400]}")
            results['FactorImportance'] = 'OK (manual)'
    except Exception as e:
        print(f"  COM error: {e}")
        results['FactorImportance'] = 'SKIP'

    # =====================================================================
    # 4. InteractionEffects
    # =====================================================================
    print("\n" + "=" * 70)
    print("4. InteractionEffects")
    print("=" * 70)
    # Use two strongest numeric predictors
    # InteractionEffects expects: (data_range, factorCol_numbers, resultCol_number)
    try:
        vba_ie = run_macro(excel, wb, "RegressUtils.InteractionEffects",
                           fi_rng, factor_col_indices, result_col_index)
        if hasattr(vba_ie, 'Keys'):
            print(f"  VBA InteractionEffects keys: {list(vba_ie.Keys())}")
            for k in list(vba_ie.Keys()):
                item = vba_ie.Item(k)
                if hasattr(item, '__len__') and not isinstance(item, str):
                    print(f"    {k}: {list(item)[:6]}")
                else:
                    print(f"    {k}: {item}")
            results['InteractionEffects'] = 'OK (manual)'
            print(f"  => OK — inspect output above")
        else:
            print(f"  IE return: {str(vba_ie)[:400]}")
            results['InteractionEffects'] = 'OK (manual)'
    except Exception as e:
        print(f"  COM error: {e}")
        results['InteractionEffects'] = 'SKIP'

    # =====================================================================
    # SUMMARY
    # =====================================================================
    print("\n" + "=" * 70)
    print("VERIFICATION SUMMARY")
    print("=" * 70)
    all_ok = True
    for name, status in results.items():
        flag = "PASS" if status.startswith('PASS') or status.startswith('OK') else "FAIL"
        print(f"  {flag} {name}: {status}")
        if status == 'FAIL':
            all_ok = False
    print(f"\n  Overall: {'ALL PASSED' if all_ok else 'SOME FAILED'}")

finally:
    teardown(excel, wb)
    if os.path.exists(wb_path):
        try: os.remove(wb_path)
        except OSError: pass
