"""Shared cross-validation runner for build_<Module>.py scripts.

Tolerance Tiers
--------------
  ========  ========  ==================================================
  Tier      tol       Used for
  ========  ========  ==================================================
  Exact     0         String comparison, boolean results
  Standard  1e-10     Basic arithmetic, sorting, array manipulation
  Numeric   1e-6      Iterative algorithms (PolyFit, matrix decompositions)
  Loose     1e-5      SVD singular values (iterative convergence)
  Stats     1e-2      Higher-order moments (skewness, kurtosis)
  Phys      1e-3..0.1 Physical constants (molecular weight, unit conv.)
  Time      1.0       Unix timestamps (1-second tolerance)
  ========  ========  ==================================================

Usage in each build_<Module>.py::

    from tests.crossval.build_common import CrossValRunner

    TEST_CASES = [
        {"name": "ArraySort_asc", "func": "ArraySort", "args": lambda: ([3,1,4,2],),
         "py_ref": lambda a: sorted(a[0]), "tol": 1e-10},
        ...
    ]

    runner = CrossValRunner("ArrayUtils", module_paths)
    results = runner.run_all(TEST_CASES)
    runner.print_summary()

Each build_*.py is also runnable standalone: ``python tests/build_ArrayUtils.py``.
"""

import os
import re
import sys
import tempfile
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple, Union

import numpy as np

# Ensure project root is on the path so relative imports resolve
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.test_utils import (  # noqa: E402
    ensure_excel,
    teardown,
    create_workbook,
    inject_testrunner,
    run_macro,
    check_close,
    write_range,
    com_to_numpy,
    PROJECT_ROOT,
    SRC_DIR,
    VBA_CORE_DIR,
    VBA_CORE_IMPORT_ORDER,
)

# Optional scipy import
try:
    from scipy import stats as _sp_stats  # noqa: F401
    _HAS_SCIPY = True
except ImportError:
    _HAS_SCIPY = False

# =============================================================================
# Test case definition types
# =============================================================================
# Each test case is a dict:
#   name       str     — unique test name within the module
#   func       str     — VBA function name (e.g., "ArraySort")
#   args       callable -> tuple — lambda that generates positional arguments
#   py_ref     callable(tuple) -> Any — computes expected Python reference value
#   tol        float   — absolute tolerance for check_close (default 1e-10)
#   result_type str   — "scalar" (default), "array", "bool", "string", "tuple"
#   skip_if    bool    — if True, skip this test
#   skip_reason str    — reason for skipping (printed in report)
#   is_udf     bool    — if True, first arg is Range data (written to sheet)
#   kwargs     dict    — keyword args for VBA function (passed after positional)
# =============================================================================


class CrossValRunner:
    """Runs VBA function tests against Python reference values via COM."""

    def __init__(
        self,
        module_name: str,
        module_paths: Sequence[str],
        extra_imports: Optional[Sequence[str]] = None,
    ):
        """Initialize the runner for a specific module.

        Parameters
        ----------
        module_name : str
            Name of the primary module under test (e.g., "ArrayUtils").
        module_paths : sequence of str
            Absolute paths to .bas/.cls files to import into the workbook.
        extra_imports : sequence of str, optional
            Additional module names (without extension) to guarantee import
            order, beyond VBA_CORE_IMPORT_ORDER.
        """
        self.module_name = module_name
        self.module_paths = list(module_paths)
        self.extra_imports = extra_imports
        self.results: List[Tuple[str, str, str, str]] = []  # (module, name, status, detail)

        # Build import order
        self._import_order = list(VBA_CORE_IMPORT_ORDER)
        if extra_imports:
            for name in extra_imports:
                if name not in self._import_order:
                    self._import_order.append(name)

    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    def run_all(self, test_cases: List[dict]) -> List[Tuple[str, str, str, str]]:
        """Run all test cases and return results."""
        excel = ensure_excel()
        wb = None
        try:
            output = os.path.join(tempfile.gettempdir(),
                                  f"vba_crossval_{self.module_name}.xlsm")
            wb = create_workbook(
                excel, output, self.module_paths,
                import_order=self._import_order,
            )
            inject_testrunner(wb)

            ws = wb.Sheets("TestData")
            self.results = []

            for tc in test_cases:
                self._run_one(excel, wb, ws, tc)

            return self.results
        finally:
            teardown(excel, wb)

    def print_summary(self) -> Tuple[int, int]:
        """Print a summary report and return (pass_count, fail_count)."""
        passed = sum(1 for r in self.results if r[2] == "PASS")
        failed = sum(1 for r in self.results if r[2] == "FAIL")
        skipped = sum(1 for r in self.results if r[2] == "SKIP")
        total = len(self.results)

        print(f"\n{'=' * 60}")
        print(f"  {self.module_name} Cross-Validation")
        print(f"{'=' * 60}")
        print(f"  Total  : {total}")
        print(f"  PASS   : {passed}")
        print(f"  FAIL   : {failed}")
        if skipped:
            print(f"  SKIP   : {skipped}")
        if total > 0:
            rate = 100.0 * (passed + skipped) / total
            print(f"  Rate   : {rate:.1f}%")
        print(f"{'=' * 60}")

        if failed > 0:
            print(f"\n  Failures ({failed}):")
            for mod, func, status, detail in self.results:
                if status == "FAIL":
                    print(f"    - {mod}.{func}: {detail}")

        return passed, failed

    # -------------------------------------------------------------------------
    # Internal helpers
    # -------------------------------------------------------------------------

    def _run_one(self, excel, wb, ws, tc: dict) -> None:
        """Execute a single test case."""
        label = f"{self.module_name}.{tc['func']}.{tc['name']}"

        # Check skip condition
        if tc.get("skip_if", False):
            print(f"  SKIP  {label} — {tc.get('skip_reason', 'no reason')}")
            self.results.append((self.module_name, tc["name"], "SKIP",
                                 tc.get("skip_reason", "")))
            return

        try:
            # Generate arguments
            args = tc["args"]() if callable(tc["args"]) else tc.get("args", ())
            if not isinstance(args, (tuple, list)):
                args = (args,)

            # Handle reconstruction tests (multi-call VBA → Python combine)
            if tc.get("reconstruct"):
                vba_result, py_val, tol = tc["reconstruct"](excel, wb, ws, self, tc, args)
            elif tc.get("range_path"):
                vba_result = self._call_vba_range(excel, wb, ws, tc, args)
                py_val = tc["py_ref"](args) if callable(tc["py_ref"]) else tc["py_ref"]
                tol = tc.get("tol", 1e-10)
            elif tc.get("is_udf"):
                vba_result = self._call_udf(excel, wb, ws, tc, args)
                py_val = tc["py_ref"](args) if callable(tc["py_ref"]) else tc["py_ref"]
                tol = tc.get("tol", 1e-10)
            else:
                vba_result = self._call_vba(excel, wb, tc, args)
                py_val = tc["py_ref"](args) if callable(tc["py_ref"]) else tc["py_ref"]
                tol = tc.get("tol", 1e-10)

            # Compare
            result_type = tc.get("result_type", "scalar")
            self._compare(label, vba_result, py_val, result_type, tol, tc)

        except Exception as exc:
            self.results.append((self.module_name, tc["name"], "FAIL",
                                f"exception: {exc}"))
            print(f"  FAIL  {label} — exception: {exc}")

    @staticmethod
    def _to_com_arg(arg):
        """Convert Python types to COM-friendly values.

        datetime/date → ISO date string (e.g. "2024-02-01").
        VBA CDate() parses this unambiguously, avoiding timezone shift
        and date-serial ambiguity.
        """
        from datetime import date as date_cls, datetime as dt_cls
        if isinstance(arg, (date_cls, dt_cls)):
            return arg.strftime("%Y-%m-%d")
        return arg

    def _call_vba(self, excel, wb, tc: dict, args: tuple) -> Any:
        """Call a VBA function directly via Application.Run."""
        macro = f"{self.module_name}.{tc['func']}"
        args = tuple(self._to_com_arg(a) for a in args)
        return run_macro(excel, wb, macro, *args)

    def _call_vba_range(self, excel, wb, ws, tc: dict, args: tuple) -> Any:
        """Call a VBA function with Range objects instead of Python lists.

        Array/list args are written to the TestData sheet and replaced with
        Range objects. Scalar args pass through as-is. This tests the
        Range→Variant COM marshaling path that is NOT covered by _call_vba.
        """
        ws.UsedRange.ClearContents()
        macro = f"{self.module_name}.{tc['func']}"
        rng_args = []
        r_offset = 1
        for a in args:
            a = self._to_com_arg(a)
            if isinstance(a, (list, tuple)):
                arr = np.asarray(a, dtype=object)
                if arr.ndim == 0:
                    arr = arr.reshape(1, 1)
                elif arr.ndim == 1:
                    arr = arr.reshape(-1, 1)
                nr, nc = arr.shape
                write_range(ws, arr, r_offset, 1)
                rng = ws.Range(ws.Cells(r_offset, 1),
                               ws.Cells(r_offset + nr - 1, nc))
                rng_args.append(rng)
                r_offset = r_offset + nr + 1  # blank row separator
            else:
                rng_args.append(a)
        return run_macro(excel, wb, macro, *rng_args)

    def _call_udf(self, excel, wb, ws, tc: dict, args: tuple) -> Any:
        """Call a UDF that takes Range input(s).

        Array/list args are written to the TestData sheet as separate ranges
        (with a blank row between them). Scalar args pass through as-is.
        """
        ws.UsedRange.ClearContents()
        rng_args = []
        r_offset = 1
        macro = tc["func"]
        for a in args:
            if isinstance(a, (list, tuple)):
                arr = np.asarray(a, dtype=object)
                if arr.ndim == 0: arr = arr.reshape(1, 1)
                elif arr.ndim == 1: arr = arr.reshape(-1, 1)
                nr, nc = arr.shape
                write_range(ws, arr, r_offset, 1)
                rng_args.append(ws.Range(ws.Cells(r_offset, 1), ws.Cells(r_offset + nr - 1, nc)))
                r_offset = r_offset + nr + 1  # blank row separator
            else:
                rng_args.append(a)
        return run_macro(excel, wb, macro, *rng_args)

    def _compare(self, label: str, vba_result: Any, py_val: Any,
                 result_type: str, tol: float, tc: dict) -> None:
        """Compare VBA and Python results based on result_type or compare_mode."""
        try:
            cmp_mode = tc.get("compare_mode", "")
            if cmp_mode == "dict_keys":
                self._cmp_dict_keys(label, vba_result, py_val, tc)
            elif cmp_mode == "dict_keys_range":
                self._cmp_dict_range(label, vba_result, py_val, tc)
            elif result_type == "array":
                self._compare_array(label, vba_result, py_val, tol)
            elif result_type == "bool":
                self._compare_bool(label, vba_result, py_val)
            elif result_type == "string":
                self._compare_string(label, vba_result, py_val)
            else:
                self._compare_scalar(label, vba_result, py_val, tol)
        except Exception as exc:
            self.results.append((self.module_name, tc["name"], "FAIL",
                                f"compare error: {exc}"))
            print(f"  FAIL  {label} — compare error: {exc}")

    def _cmp_dict_keys(self, label, vba_obj, expected_keys, tc):
        """Verify VBA Dictionary has expected keys (COM .Exists check)."""
        if vba_obj is None:
            self.results.append((self.module_name, tc["name"], "FAIL", "null dict"))
            print(f"  FAIL  {label} — null result"); return
        missing = [k for k in expected_keys if not vba_obj.Exists(k)]
        if missing:
            self.results.append((self.module_name, tc["name"], "FAIL",
                                 f"missing keys: {missing}"))
            print(f"  FAIL  {label} — missing keys: {missing}")
        else:
            self.results.append((self.module_name, tc["name"], "PASS", ""))
            print(f"  PASS  {label}  all {len(expected_keys)} keys present")

    def _cmp_dict_range(self, label, vba_obj, checks, tc):
        """Verify Dictionary key values within ranges. checks=[(key, lo, hi),...]"""
        if vba_obj is None:
            self.results.append((self.module_name, tc["name"], "FAIL", "null dict"))
            print(f"  FAIL  {label} — null result"); return
        failures = []
        for key, lo, hi in checks:
            val = float(vba_obj(key))
            if not (lo <= val <= hi):
                failures.append(f"{key}={val:.4g} not in [{lo}, {hi}]")
        if failures:
            self.results.append((self.module_name, tc["name"], "FAIL",
                                 "; ".join(failures)))
            print(f"  FAIL  {label} — {'; '.join(failures)}")
        else:
            self.results.append((self.module_name, tc["name"], "PASS", ""))
            print(f"  PASS  {label}  all {len(checks)} ranges OK")

    def _compare_scalar(self, label, vba_result, py_val, tol):
        """Compare scalar (float/int/date) results."""
        # Handle datetime/date comparisons (VBA Date returns via COM)
        from datetime import date as date_cls, datetime as dt_cls
        if isinstance(vba_result, (date_cls, dt_cls)):
            vba_dt = dt_cls(vba_result.year, vba_result.month, vba_result.day)
            py_dt = dt_cls(py_val.year, py_val.month, py_val.day) if isinstance(py_val, (date_cls, dt_cls)) else None
            if py_dt is not None:
                diff_days = abs((vba_dt - py_dt).total_seconds() / 86400.0)
                if diff_days < 1.0:
                    self.results.append((self.module_name, label, "PASS", ""))
                    print(f"  PASS  {label}  VBA={vba_dt.date()}  PY={py_dt.date()}")
                else:
                    self.results.append((self.module_name, label, "FAIL",
                                        f"date mismatch: VBA={vba_dt.date()} PY={py_dt.date()}"))
                    print(f"  FAIL  {label} — date mismatch: VBA={vba_dt.date()} PY={py_dt.date()}")
                return
            # py_val is not a date — convert VBA date to serial for numeric comparison
            epoch = dt_cls(1899, 12, 30)
            vba_val = (vba_dt - epoch).total_seconds() / 86400.0
        else:
            vba_val = _com_to_scalar(vba_result)

        if vba_val is None:
            self.results.append((self.module_name, label, "FAIL", "null result"))
            print(f"  FAIL  {label} — null result")
            return

        status = check_close(label, vba_val, py_val, tol)
        if status is True:
            self.results.append((self.module_name, label, "PASS", ""))
            print(f"  PASS  {label}  VBA={vba_val!r}  PY={py_val!r}")
        else:
            self.results.append((self.module_name, label, "FAIL", str(status)))
            print(f"  FAIL  {label} — {status}")

    def _compare_array(self, label, vba_result, py_val, tol):
        """Compare array results element-wise. Handles both numeric and string arrays."""
        # Check if this is a string array
        is_str = False
        if isinstance(vba_result, (tuple, list)) and len(vba_result) > 0:
            first = vba_result[0]
            if isinstance(first, (tuple, list)) and len(first) > 0:
                first = first[0]
            if isinstance(first, str):
                try: float(first)
                except (ValueError, TypeError): is_str = True

        if is_str:
            vba_flat = _flatten_com_result(vba_result)
            if isinstance(py_val, (list, tuple)):
                py_str = ",".join(str(x) for x in _flatten_list(py_val))
            else:
                py_str = str(py_val)
            ok = re.sub(r'\s+', ' ', vba_flat).strip() == re.sub(r'\s+', ' ', py_str).strip()
            s = "PASS" if ok else "FAIL"
            d = "" if ok else f"VBA={vba_flat!r} PY={py_str!r}"
            self.results.append((self.module_name, label, s, d))
            try:
                print(f"  {s:4s} {label}  VBA={vba_flat!r}  PY={py_str!r}")
            except UnicodeEncodeError:
                if ok:
                    print(f"  {s:4s} {label}  (string array comparison ok)")
                else:
                    print(f"  {s:4s} {label} — {d}")
            return

        vba_arr = com_to_numpy(vba_result)
        if isinstance(py_val, (list, tuple)):
            py_val = np.asarray(py_val, dtype=float)
        elif isinstance(py_val, np.ndarray):
            pass
        else:
            py_val = np.asarray([py_val], dtype=float)

        # Reshape for broadcasting comparison
        if py_val.ndim == 1 and vba_arr.ndim == 2 and vba_arr.shape[1] == 1:
            py_val = py_val.reshape(-1, 1)
        elif py_val.ndim == 2 and vba_arr.ndim == 1:
            vba_arr = vba_arr.reshape(-1, 1)

        # Handle empty arrays: np.max on empty raises ValueError
        if vba_arr.size == 0 and py_val.size == 0:
            max_err = 0.0
            ok = True
        elif vba_arr.size == 0 or py_val.size == 0:
            ok = False
            max_err = float("inf")
        else:
            try:
                max_err = float(np.max(np.abs(vba_arr - py_val)))
                ok = max_err <= tol
            except (ValueError, TypeError):
                # Shape mismatch or non-numeric — fall back to string compare
                vba_str = str(vba_result)
                py_str = str(py_val)
                ok = vba_str == py_str
                max_err = float("inf") if not ok else 0.0

        if ok:
            self.results.append((self.module_name, label, "PASS", ""))
            print(f"  PASS  {label}  max|err|={max_err:.2e}")
        else:
            self.results.append((self.module_name, label, "FAIL",
                                f"max|err|={max_err:.2e} > tol={tol}"))
            print(f"  FAIL  {label} — max|err|={max_err:.2e} > tol={tol}")

    def _compare_bool(self, label, vba_result, py_val):
        """Compare boolean results (VBA returns -1 for True, 0 for False)."""
        if vba_result is None:
            self.results.append((self.module_name, label, "FAIL", "null result"))
            print(f"  FAIL  {label} — null result")
            return
        try:
            vba_bool = bool(int(float(vba_result)))
        except (ValueError, TypeError):
            vba_bool = str(vba_result).strip().upper() == "TRUE"
        py_bool = bool(py_val)
        ok = vba_bool == py_bool
        status = "PASS" if ok else "FAIL"
        detail = "" if ok else f"VBA={vba_bool!r} PY={py_bool!r}"
        self.results.append((self.module_name, label, status, detail))
        if ok:
            print(f"  PASS  {label}  VBA={vba_bool}  PY={py_bool}")
        else:
            print(f"  FAIL  {label} — {detail}")

    def _compare_string(self, label, vba_result, py_val):
        """Compare string results."""
        vba_str = str(vba_result).strip() if vba_result is not None else ""
        py_str = str(py_val).strip()
        ok = vba_str == py_str
        status = "PASS" if ok else "FAIL"
        detail = "" if ok else f"VBA={vba_str!r} PY={py_str!r}"
        self.results.append((self.module_name, label, status, detail))
        try:
            if ok:
                print(f"  PASS  {label}  VBA={vba_str!r}  PY={py_str!r}")
            else:
                print(f"  FAIL  {label} — {detail}")
        except UnicodeEncodeError:
            # Windows console (e.g. GBK) may reject Unicode characters
            if ok:
                print(f"  PASS  {label}  (string comparison ok)")
            else:
                print(f"  FAIL  {label} — {status}")


# =============================================================================
# Utility functions
# =============================================================================

def _com_to_scalar(vba_result: Any) -> Optional[float]:
    """Extract a single float from a COM return value."""
    if vba_result is None:
        return None
    if isinstance(vba_result, (int, float)):
        return float(vba_result)
    if isinstance(vba_result, (tuple, list)):
        if len(vba_result) == 0:
            return None
        if len(vba_result) == 1:
            inner = vba_result[0]
            if isinstance(inner, (tuple, list)):
                if len(inner) == 1:
                    return float(inner[0])
                elif len(inner) > 0:
                    return float(inner[0])
            return float(inner)
        # First element of first row
        if isinstance(vba_result[0], (tuple, list)):
            return float(vba_result[0][0])
        return float(vba_result[0])
    try:
        return float(vba_result)
    except (ValueError, TypeError):
        return None


def _flatten_com_result(val) -> str:
    """Flatten a nested COM tuple result to comma-separated string."""
    if isinstance(val, (tuple, list)):
        parts = []
        for item in val:
            if isinstance(item, (tuple, list)):
                parts.append(_flatten_com_result(item))
            else:
                parts.append(str(item))
        return ",".join(parts)
    return str(val)


def _flatten_list(val) -> list:
    """Flatten a nested list/tuple to a flat list."""
    result = []
    for item in val:
        if isinstance(item, (list, tuple)):
            result.extend(_flatten_list(item))
        else:
            result.append(item)
    return result


def load_crossval_data() -> Dict[str, Dict[str, np.ndarray]]:
    """Load the cross-validation reference workbook.

    Returns a dict keyed by sheet name, each value a dict keyed by column
    header -> 1-D numpy array of floats (NaN for missing/non-numeric).
    """
    import openpyxl

    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "data", "Cross_Validation_vs_Python.xlsx",
    )
    data: Dict[str, Dict[str, np.ndarray]] = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2 or ws.max_column < 1:
                continue
            headers = [
                str(ws.cell(1, c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            sheet_data: Dict[str, List[float]] = {h: [] for h in headers if h}
            for r in range(2, ws.max_row + 1):
                for c, h in enumerate(headers, start=1):
                    if not h:
                        continue
                    raw = ws.cell(r, c).value
                    if raw is None:
                        sheet_data[h].append(np.nan)
                    elif isinstance(raw, (int, float)):
                        sheet_data[h].append(float(raw))
                    else:
                        try:
                            sheet_data[h].append(float(raw))
                        except (ValueError, TypeError):
                            sheet_data[h].append(np.nan)
            for h in sheet_data:
                sheet_data[h] = np.array(sheet_data[h], dtype=float)
            data[sheet_name] = sheet_data
    finally:
        wb.close()
    return data


def load_text_data() -> Dict[str, Dict[str, List[str]]]:
    """Load text columns from the cross-validation reference workbook.

    Returns a dict keyed by sheet name, each value a dict keyed by column
    header -> list of strings.
    """
    import openpyxl

    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "data", "Cross_Validation_vs_Python.xlsx",
    )
    data: Dict[str, Dict[str, List[str]]] = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        for sheet_name in ["TextData", "TextData1"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if ws.max_row < 2 or ws.max_column < 1:
                continue
            headers = [
                str(ws.cell(1, c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            sheet_data: Dict[str, List[str]] = {h: [] for h in headers if h}
            for r in range(2, ws.max_row + 1):
                for c, h in enumerate(headers, start=1):
                    if not h:
                        continue
                    raw = ws.cell(r, c).value
                    sheet_data[h].append(str(raw) if raw is not None else "")
            data[sheet_name] = sheet_data
    finally:
        wb.close()
    return data


# =============================================================================
# Standalone self-test
# =============================================================================
if __name__ == "__main__":
    print("=== build_common self-check ===")

    # Test _com_to_scalar
    assert _com_to_scalar(3.14) == 3.14
    assert _com_to_scalar(42) == 42.0
    assert _com_to_scalar(((1.0,),)) == 1.0
    assert _com_to_scalar(None) is None
    assert _com_to_scalar(()) is None
    print("  _com_to_scalar  : OK")

    # Test load_crossval_data
    try:
        data = load_crossval_data()
        assert "SourceData" in data
        assert "Temperature" in data["SourceData"]
        assert len(data["SourceData"]["Temperature"]) > 0
        print(f"  load_crossval_data : OK ({len(data)} sheets)")
    except FileNotFoundError:
        print("  load_crossval_data : SKIP (data file not found)")

    # Test load_text_data
    try:
        text_data = load_text_data()
        assert "TextData" in text_data
        print(f"  load_text_data   : OK ({len(text_data)} sheets)")
    except FileNotFoundError:
        print("  load_text_data   : SKIP (data file not found)")

    # Test CrossValRunner instantiation
    runner = CrossValRunner("TestModule", [])
    assert runner.module_name == "TestModule"
    print("  CrossValRunner   : OK")

    print("\n  All self-checks passed.")
